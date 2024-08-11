import openpyxl

def getrowCount(file,sheet_name):
    book = openpyxl.load_workbook(file) #file from which we have to read the data
    sheet = book[sheet_name] # sheet from which we have to read the data
    return sheet.max_row

def readData(file, sheet_name, rowno, colno):
    book = openpyxl.load_workbook(file)
    sheet = book[sheet_name]
    return sheet.cell(row=rowno,column=colno).value

def writeData(file, sheet_name, rowno, colno, data):
    book = openpyxl.load_workbook(file)
    sheet = book[sheet_name]
    sheet.cell(row=rowno,column=colno).value = data
    book.save(file)

